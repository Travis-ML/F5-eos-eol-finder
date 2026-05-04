"""Annotate an uploaded .xlsx workbook with F5 lifecycle data.

Strategy: walk every cell in every sheet. For each row that contains at least
one F5 SKU, write the matched lifecycle data into new columns to the right of
the existing data. A summary sheet is inserted at the front of the workbook.
"""

from __future__ import annotations

import io
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from matcher import Match, data_revision, match


F5_SKU_LIKE = re.compile(r"^\s*F5[\s-]", re.IGNORECASE)

OUTPUT_HEADERS = [
    "F5 EoL — Family / Category",
    "Status",
    "End of Sale",
    "End of Software Dev",
    "End of Technical Support",
    "End of RMA",
    "Notes",
]

STATUS_LABEL = {
    "regular": "Regular Support",
    "regular_no_eos": "Regular Support (no EoS announced)",
    "eos_announced": "EoS announced",
    "eosd": "Past End of Software Dev",
    "eots": "Past End of Technical Support",
}

STATUS_FILL = {
    "regular": PatternFill("solid", fgColor="C6EFCE"),
    "regular_no_eos": PatternFill("solid", fgColor="C6EFCE"),
    "eos_announced": PatternFill("solid", fgColor="FFEB9C"),
    "eosd": PatternFill("solid", fgColor="FFC7CE"),
    "eots": PatternFill("solid", fgColor="FF9999"),
}


@dataclass
class Stats:
    rows_annotated: int = 0
    by_status: Counter = None
    by_category: Counter = None
    unknown_skus: set = None

    def __post_init__(self):
        if self.by_status is None:
            self.by_status = Counter()
        if self.by_category is None:
            self.by_category = Counter()
        if self.unknown_skus is None:
            self.unknown_skus = set()


def _looks_like_f5_sku(v: object) -> bool:
    return isinstance(v, str) and bool(F5_SKU_LIKE.match(v))


def _pick_best_match(matches: list[Match]) -> Optional[Match]:
    """When a row has multiple F5 SKUs (e.g. Part Number + Covered Product),
    prefer hardware matches over non-hardware over unknown."""
    if not matches:
        return None
    order = {"hardware": 0, "non_hardware": 1, "unknown": 2, "non_f5": 3}
    return sorted(matches, key=lambda m: order.get(m.kind, 9))[0]


def _format_date(d) -> str:
    if d is None:
        return ""
    if isinstance(d, date):
        return d.strftime("%Y-%m-%d")
    return str(d)


def _row_values(m: Match) -> list:
    if m.kind == "hardware":
        return [
            m.display_name,
            STATUS_LABEL.get(m.status, m.status or ""),
            _format_date(m.end_of_sale),
            _format_date(m.end_of_software_dev),
            _format_date(m.end_of_technical_support),
            _format_date(m.end_of_rma),
            m.note or "",
        ]
    if m.kind == "non_hardware":
        return [m.category, "n/a", "", "", "", "", m.note or ""]
    if m.kind == "unknown":
        return [
            "Unknown F5 SKU",
            "needs review",
            "",
            "",
            "",
            "",
            "SKU starts with F5- but is not in the lifecycle DB. Verify manually.",
        ]
    return ["", "", "", "", "", "", ""]


def annotate_workbook(file_bytes: bytes) -> tuple[bytes, Stats]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    stats = Stats()

    for ws in list(wb.worksheets):
        max_col_before = ws.max_column
        out_start = max_col_before + 2  # leave one blank column for readability

        header_row = _detect_header_row(ws)
        if header_row is not None:
            for i, h in enumerate(OUTPUT_HEADERS):
                cell = ws.cell(row=header_row, column=out_start + i, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=max_col_before):
            row_matches = []
            target_row_idx = None
            for cell in row:
                if _looks_like_f5_sku(cell.value):
                    target_row_idx = cell.row
                    row_matches.append(match(cell.value))
            if not row_matches:
                continue

            picked = _pick_best_match(row_matches)
            stats.rows_annotated += 1
            if picked.kind == "hardware":
                stats.by_status[picked.status or "unknown"] += 1
            elif picked.kind == "non_hardware":
                stats.by_category[picked.category or "Other"] += 1
            elif picked.kind == "unknown":
                stats.unknown_skus.add(picked.sku)

            values = _row_values(picked)
            fill = STATUS_FILL.get(picked.status) if picked.kind == "hardware" else None
            for i, v in enumerate(values):
                c = ws.cell(row=target_row_idx, column=out_start + i, value=v)
                if fill:
                    c.fill = fill

        for i in range(len(OUTPUT_HEADERS)):
            ws.column_dimensions[get_column_letter(out_start + i)].width = 22

    _insert_summary_sheet(wb, stats)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), stats


def _detect_header_row(ws) -> Optional[int]:
    """Find the row index of a likely header by looking for cells containing
    a known SKU column label. Falls back to None if not found (in which case
    we just don't write headers, but data is still annotated)."""
    needles = {"product sku", "part number", "covered product", "sku"}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() in needles:
                return cell.row
    return None


def _insert_summary_sheet(wb, stats: Stats) -> None:
    summary = wb.create_sheet("F5 EoL Summary", 0)
    bold = Font(bold=True)

    summary["A1"] = "F5 End-of-Life / End-of-Support Report"
    summary["A1"].font = Font(bold=True, size=14)

    summary["A3"] = "Lifecycle data revision:"
    summary["B3"] = data_revision() or "unknown"
    summary["A3"].font = bold

    summary["A4"] = "Rows annotated:"
    summary["B4"] = stats.rows_annotated
    summary["A4"].font = bold

    row = 6
    summary.cell(row=row, column=1, value="Hardware by lifecycle status").font = bold
    row += 1
    if stats.by_status:
        for status, n in sorted(stats.by_status.items(), key=lambda x: -x[1]):
            summary.cell(row=row, column=1, value=STATUS_LABEL.get(status, status))
            summary.cell(row=row, column=2, value=n)
            fill = STATUS_FILL.get(status)
            if fill:
                summary.cell(row=row, column=1).fill = fill
            row += 1
    else:
        summary.cell(row=row, column=1, value="(none)")
        row += 1

    row += 1
    summary.cell(row=row, column=1, value="Non-hardware lines by category").font = bold
    row += 1
    if stats.by_category:
        for cat, n in sorted(stats.by_category.items(), key=lambda x: -x[1]):
            summary.cell(row=row, column=1, value=cat)
            summary.cell(row=row, column=2, value=n)
            row += 1
    else:
        summary.cell(row=row, column=1, value="(none)")
        row += 1

    row += 1
    summary.cell(row=row, column=1, value="Unknown F5 SKUs (need manual review)").font = bold
    row += 1
    if stats.unknown_skus:
        for sku in sorted(stats.unknown_skus):
            summary.cell(row=row, column=1, value=sku)
            row += 1
    else:
        summary.cell(row=row, column=1, value="(none)")

    summary.column_dimensions["A"].width = 50
    summary.column_dimensions["B"].width = 12

    row += 2
    summary.cell(row=row, column=1, value=(
        "Source: F5 K4309 Hardware Product Lifecycle Support Policy. "
        "Verify against my.f5.com/manage/s/article/K4309 before relying on these dates."
    )).alignment = Alignment(wrap_text=True)
